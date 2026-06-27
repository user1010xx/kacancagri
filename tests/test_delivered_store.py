from datetime import date, datetime
from pathlib import Path

from delivered_store import DeliveredStore


def test_add_and_get_by_call_date(tmp_path: Path):
    store = DeliveredStore(tmp_path / "delivered.json")
    d = date(2026, 6, 27)
    store.add(
        call_key="k1",
        phone="905301718596",
        personel_adi="Seda",
        call_date=d,
        notified_at=datetime(2026, 6, 27, 11, 5, 0),
    )
    rows = store.get_by_call_date(d)
    assert len(rows) == 1
    assert rows[0]["phone"] == "905301718596"
    assert rows[0]["personel_adi"] == "Seda"
    assert rows[0]["notified_at"] == "27.06.2026 11:05:00"


def test_dedup_by_call_key(tmp_path: Path):
    store = DeliveredStore(tmp_path / "delivered.json")
    d = date(2026, 6, 27)
    store.add(call_key="k1", phone="1", personel_adi="A", call_date=d)
    store.add(call_key="k1", phone="1", personel_adi="A", call_date=d)
    assert store.count() == 1


def test_export_delivered_report_excel(tmp_path: Path):
    from excel_export import export_delivered_report_excel
    from openpyxl import load_workbook

    rows = [
        {
            "phone": "905301718596",
            "personel_adi": "Seda",
            "notified_at": "27.06.2026 11:05:00",
        }
    ]
    path = export_delivered_report_excel(rows, tmp_path / "rapor.xlsx")
    wb = load_workbook(path)
    ws = wb.active
    assert ws.cell(1, 1).value == "Numara"
    assert ws.cell(2, 1).value == "905301718596"
    assert ws.cell(2, 2).value == "Seda"
    assert ws.cell(2, 3).value == "27.06.2026 11:05:00"


def test_purge_call_date(tmp_path: Path):
    store = DeliveredStore(tmp_path / "delivered.json")
    d1 = date(2026, 6, 26)
    d2 = date(2026, 6, 27)
    store.add(call_key="k1", phone="1", personel_adi="A", call_date=d1)
    store.add(call_key="k2", phone="2", personel_adi="B", call_date=d2)
    removed = store.purge_call_date(d1)
    assert removed == 1
    assert store.count() == 1