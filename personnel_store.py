import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class PersonnelStore:
    """Personel yönetimi: dahili_ad -> {personel_adi, telegram_username}
    Excel ile güncelleme destekler (mevcutları günceller, yenileri ekler).
    """

    def __init__(self, path: Path) -> None:
        self.path = path
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self._data: dict[str, dict[str, str]] = {}
        self._load()

    def _load(self) -> None:
        if self.path.exists():
            try:
                with self.path.open("r", encoding="utf-8") as f:
                    loaded = json.load(f)
                if isinstance(loaded, dict):
                    self._data = loaded
            except Exception:
                self._data = {}

    def _save(self) -> None:
        with self.path.open("w", encoding="utf-8") as f:
            json.dump(self._data, f, ensure_ascii=False, indent=2)

    def add_or_update(self, dahili_ad: str, personel_adi: str, telegram_username: str) -> bool:
        dahili = str(dahili_ad).strip()
        if not dahili:
            return False

        self._data[dahili] = {
            "personel_adi": str(personel_adi).strip(),
            "telegram_username": str(telegram_username).strip().lstrip("@"),
        }
        self._save()
        return True

    def remove(self, dahili_ad: str) -> bool:
        dahili = str(dahili_ad).strip()
        if dahili in self._data:
            del self._data[dahili]
            self._save()
            return True
        return False

    def get(self, dahili_ad: str) -> dict[str, str] | None:
        return self._data.get(str(dahili_ad).strip())

    def get_all(self) -> list[dict[str, str]]:
        result = []
        for dahili, info in self._data.items():
            result.append(
                {
                    "dahili_ad": dahili,
                    "personel_adi": info.get("personel_adi", ""),
                    "telegram_username": info.get("telegram_username", ""),
                }
            )
        # Dahili adına göre sırala
        return sorted(result, key=lambda x: x["dahili_ad"])

    def load_from_excel(self, excel_path: Path) -> int:
        """3 sütunluk Excel'i okur: dahili_ad, personel_adi, telegram_username.
        Mevcutları günceller, yenileri ekler. İşlenen satır sayısını döner.
        """
        if not excel_path.exists():
            return 0

        count = 0
        wb = load_workbook(excel_path, read_only=True)
        try:
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                if not row:
                    continue
                # İlk 3 sütunu al
                dahili = str(row[0]).strip() if row[0] else ""
                ad = str(row[1]).strip() if row[1] else ""
                username = str(row[2]).strip() if len(row) > 2 and row[2] else ""

                # Header satırlarını atla
                if not dahili or dahili.lower() in {"dahili", "dahili_ad", "extension", "extensionname"}:
                    continue
                if not ad:
                    continue

                if self.add_or_update(dahili, ad, username):
                    count += 1
        finally:
            wb.close()

        return count

    def count(self) -> int:
        return len(self._data)